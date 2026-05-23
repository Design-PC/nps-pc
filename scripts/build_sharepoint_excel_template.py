from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


OUTPUT = Path("deliverables/Prime_Control_Respostas_NPS_2026_Modelo_SharePoint.xlsx")

BLUE = "003F7D"
BLUE_DARK = "0E243D"
ORANGE = "FF7A21"
LIGHT_BLUE = "EAF3FB"
LIGHT_GRAY = "F3F6F8"
MID_GRAY = "D9E2EA"
TEXT = "172B4D"
WHITE = "FFFFFF"
GREEN = "D8F3DC"
RED = "FDE2E1"
YELLOW = "FFF3CD"


BASE_HEADERS = [
    "ID resposta",
    "Data de resposta",
    "Campanha",
    "Status",
    "Fonte",
    "Token / sessão",
    "Empresa",
    "Nome",
    "E-mail corporativo",
    "Cargo",
    "Responsável Prime Control",
    "Início",
    "Conclusão",
    "Tempo de conclusão (min)",
    "Dispositivo",
    "Navegador",
    "Nota NPS (1 a 10)",
    "Classificação NPS",
    "Motivo da nota",
    "Entendimento do negócio",
    "Relevância das soluções entregues",
    "Valor percebido",
    "Comprometimento com resultados",
    "Engajamento na solução de problemas",
    "Qualidade das entregas",
    "Cumprimento dos prazos",
    "Clareza e objetividade",
    "Tempo de resposta da equipe",
    "Qualidade do atendimento",
    "Empresa inovadora e alinhada ao mercado",
    "Antecipar tendências e propor soluções",
    "Áreas para investir mais em inovação",
    "Expectativa sobre parceria estratégica",
    "Iniciativas/melhorias para gerar mais valor",
    "Perguntas respondidas",
    "Campos abertos preenchidos",
    "Etapa abandonada",
    "Sinal de risco",
    "Status de follow-up",
    "Próxima ação",
    "Responsável follow-up",
    "Data follow-up",
    "Observações internas",
    "Link replay Clarity (opcional)",
    "Última atualização",
]


DICTIONARY_ROWS = [
    ("ID resposta", "Identificador único da resposta.", "Sistema", "Não"),
    ("Data de resposta", "Data/hora de conclusão da pesquisa.", "Sistema", "Não"),
    ("Campanha", "Nome da campanha vigente.", "Sistema", "Não"),
    ("Status", "Convidado, Iniciado, Em andamento, Concluído ou Abandonado.", "Sistema", "Não"),
    ("Fonte", "Origem do acesso: e-mail, lembrete, manual ou outro canal.", "Sistema", "Não"),
    ("Token / sessão", "Identificador técnico da sessão, sem expor senha ou login.", "Sistema", "Não"),
    ("Empresa", "Cliente respondente.", "Pesquisa", "Não"),
    ("Nome", "Nome do contato respondente.", "Pesquisa", "Não"),
    ("E-mail corporativo", "E-mail corporativo validado no formulário.", "Pesquisa", "Não"),
    ("Cargo", "Cargo informado pelo respondente.", "Pesquisa", "Não"),
    ("Responsável Prime Control", "Pessoa interna responsável pelo relacionamento.", "Manual interno", "Sim"),
    ("Nota NPS (1 a 10)", "Nota da pergunta de recomendação.", "Pesquisa", "Não"),
    ("Classificação NPS", "Promotor, Neutro ou Detrator, calculado pela nota.", "Sistema", "Não"),
    ("Motivo da nota", "Resposta aberta sobre o motivo da nota.", "Pesquisa", "Não"),
    ("Perguntas respondidas", "Quantidade de respostas preenchidas.", "Sistema", "Não"),
    ("Campos abertos preenchidos", "Quantidade de comentários abertos preenchidos.", "Sistema", "Não"),
    ("Etapa abandonada", "Última etapa registrada em caso de abandono.", "Sistema", "Não"),
    ("Sinal de risco", "Normal, Silencioso, Médio ou Alto.", "Sistema", "Não"),
    ("Status de follow-up", "Pendente, Em andamento, Concluído ou Não se aplica.", "Manual interno", "Sim"),
    ("Próxima ação", "Ação definida pelo time responsável.", "Manual interno", "Sim"),
    ("Link replay Clarity (opcional)", "Link manual para replay/heatmap quando houver análise no Clarity.", "Manual interno", "Sim"),
]


def style_title(cell, size=22):
    cell.font = Font(name="Calibri", bold=True, color=BLUE_DARK, size=size)
    cell.alignment = Alignment(vertical="center")


def style_section(cell):
    cell.fill = PatternFill("solid", fgColor=BLUE)
    cell.font = Font(name="Calibri", bold=True, color=WHITE, size=12)
    cell.alignment = Alignment(horizontal="center", vertical="center")


def style_header(row):
    for cell in row:
        cell.fill = PatternFill("solid", fgColor=BLUE)
        cell.font = Font(name="Calibri", bold=True, color=WHITE, size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(bottom=Side(style="thin", color=ORANGE))


def style_table(ws, min_row, max_row, min_col, max_col):
    thin = Side(style="thin", color=MID_GRAY)
    for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for cell in row:
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            if cell.row > min_row:
                cell.font = Font(name="Calibri", size=10, color=TEXT)


def set_widths(ws, widths):
    for column, width in widths.items():
        ws.column_dimensions[column].width = width


def create_workbook():
    wb = Workbook()
    ws_dashboard = wb.active
    ws_dashboard.title = "Resumo executivo"
    ws_base = wb.create_sheet("Base automática")
    ws_config = wb.create_sheet("Configurações")
    ws_dict = wb.create_sheet("Dicionário")

    build_base(ws_base)
    build_dashboard(ws_dashboard)
    build_config(ws_config)
    build_dictionary(ws_dict)

    wb.properties.title = "Prime Control - Respostas NPS 2026"
    wb.properties.subject = "Modelo de planilha oficial para SharePoint/OneDrive"
    wb.properties.creator = "Prime Control"
    wb.properties.company = "Prime Control"
    wb.save(OUTPUT)


def build_base(ws):
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(BASE_HEADERS))}1"
    ws.append(BASE_HEADERS)
    ws.append(["" for _ in BASE_HEADERS])
    style_header(ws[1])
    ws.row_dimensions[1].height = 42
    ws.row_dimensions[2].height = 24
    style_table(ws, 1, 2, 1, len(BASE_HEADERS))
    table_ref = f"A1:{get_column_letter(len(BASE_HEADERS))}2"
    table = Table(displayName="TabelaRespostasNPS", ref=table_ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)

    widths = {
        "A": 18, "B": 18, "C": 22, "D": 16, "E": 14, "F": 22, "G": 26, "H": 24,
        "I": 28, "J": 22, "K": 26, "L": 18, "M": 18, "N": 18, "O": 16, "P": 16,
        "Q": 16, "R": 18, "S": 38, "T": 18, "U": 22, "V": 18, "W": 22, "X": 22,
        "Y": 18, "Z": 18, "AA": 20, "AB": 20, "AC": 20, "AD": 24, "AE": 26,
        "AF": 34, "AG": 34, "AH": 34, "AI": 18, "AJ": 20, "AK": 18, "AL": 16,
        "AM": 20, "AN": 32, "AO": 24, "AP": 18, "AQ": 34, "AR": 32, "AS": 18,
    }
    set_widths(ws, widths)

    validations = {
        "D2:D5000": '"Convidado,Iniciado,Em andamento,Concluído,Abandonado"',
        "Q2:Q5000": '"1,2,3,4,5,6,7,8,9,10"',
        "R2:R5000": '"Promotor,Neutro,Detrator"',
        "AL2:AL5000": '"Normal,Silencioso,Médio,Alto"',
        "AM2:AM5000": '"Pendente,Em andamento,Concluído,Não se aplica"',
    }
    for rng, formula in validations.items():
        dv = DataValidation(type="list", formula1=formula, allow_blank=True)
        ws.add_data_validation(dv)
        dv.add(rng)

    ws.conditional_formatting.add(
        "Q2:Q5000",
        CellIsRule(operator="lessThanOrEqual", formula=["6"], fill=PatternFill("solid", fgColor=RED)),
    )
    ws.conditional_formatting.add(
        "Q2:Q5000",
        CellIsRule(operator="between", formula=["7", "8"], fill=PatternFill("solid", fgColor=YELLOW)),
    )
    ws.conditional_formatting.add(
        "Q2:Q5000",
        CellIsRule(operator="greaterThanOrEqual", formula=["9"], fill=PatternFill("solid", fgColor=GREEN)),
    )


def build_dashboard(ws):
    ws.sheet_view.showGridLines = False
    ws.merge_cells("A1:H1")
    ws["A1"] = "Prime Control | Respostas NPS 2026"
    style_title(ws["A1"], 24)
    ws["A2"] = "Resumo executivo conectado à aba Base automática. A plataforma deve atualizar somente a base; análises e fórmulas permanecem preservadas."
    ws.merge_cells("A2:H2")
    ws["A2"].font = Font(name="Calibri", color="52616F", size=11)

    kpis = [
        ("Total respostas", '=COUNTA(\'Base automática\'!A2:A5000)'),
        ("Concluídas", '=COUNTIF(\'Base automática\'!D:D,"Concluído")'),
        ("Taxa conclusão", '=IFERROR(C5/C4,0)'),
        ("NPS final", '=IF(C4=0,"",ROUND(((COUNTIFS(\'Base automática\'!Q:Q,">=9")-COUNTIFS(\'Base automática\'!Q:Q,"<=6"))/COUNT(\'Base automática\'!Q:Q))*100,0))'),
        ("Média nota NPS", '=IFERROR(AVERAGE(\'Base automática\'!Q:Q),"")'),
        ("Clientes de risco", '=COUNTIF(\'Base automática\'!AL:AL,"Alto")'),
    ]
    start_row = 4
    for idx, (label, formula) in enumerate(kpis):
        col = 1 + (idx % 3) * 3
        row = start_row + (idx // 3) * 3
        ws.cell(row=row, column=col, value=label)
        ws.cell(row=row + 1, column=col, value=formula)
        ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + 1)
        ws.merge_cells(start_row=row + 1, start_column=col, end_row=row + 1, end_column=col + 1)
        label_cell = ws.cell(row=row, column=col)
        value_cell = ws.cell(row=row + 1, column=col)
        label_cell.fill = PatternFill("solid", fgColor=LIGHT_BLUE)
        label_cell.font = Font(name="Calibri", bold=True, color=BLUE, size=10)
        value_cell.fill = PatternFill("solid", fgColor=WHITE)
        value_cell.font = Font(name="Calibri", bold=True, color=BLUE_DARK, size=22)
        value_cell.alignment = Alignment(horizontal="left", vertical="center")
    ws["G5"].number_format = "0%"
    ws["A8"].number_format = "0.0"

    ws["A11"] = "Distribuição NPS oficial"
    style_section(ws["A11"])
    ws.merge_cells("A11:D11")
    nps_table = [
        ("Grupo", "Faixa", "Qtd.", "% da base"),
        ("Promotores", "Notas 9-10", '=COUNTIFS(\'Base automática\'!Q:Q,">=9")', '=IFERROR(C13/SUM($C$13:$C$15),0)'),
        ("Neutros", "Notas 7-8", '=COUNTIFS(\'Base automática\'!Q:Q,">=7",\'Base automática\'!Q:Q,"<=8")', '=IFERROR(C14/SUM($C$13:$C$15),0)'),
        ("Detratores", "Notas 1-6", '=COUNTIFS(\'Base automática\'!Q:Q,"<=6")', '=IFERROR(C15/SUM($C$13:$C$15),0)'),
        ("Fórmula", "Promotores - Detratores", "", '=IFERROR(D13-D15,0)'),
    ]
    for r, row in enumerate(nps_table, 12):
        for c, value in enumerate(row, 1):
            ws.cell(r, c, value)
    style_header(ws[12])
    style_table(ws, 12, 16, 1, 4)
    for row in range(13, 17):
        ws.cell(row, 4).number_format = "0%"

    ws["F11"] = "Leitura temática | 25% por bloco"
    style_section(ws["F11"])
    ws.merge_cells("F11:H11")
    themes = [
        ("Relacionamento e Satisfação", "25%", '=IFERROR(AVERAGE(\'Base automática\'!Q:Q),"")'),
        ("Percepção de Valor", "25%", '=IFERROR(SUM(\'Base automática\'!T:X)/COUNT(\'Base automática\'!T:X),"")'),
        ("Qualidade Operacional", "25%", '=IFERROR(SUM(\'Base automática\'!Y:AC)/COUNT(\'Base automática\'!Y:AC),"")'),
        ("Inovação, Transformação e Futuro", "25%", '=IFERROR(SUM(\'Base automática\'!AD:AE)/COUNT(\'Base automática\'!AD:AE),"")'),
    ]
    ws.append([])
    for r, row in enumerate([("Bloco", "Peso", "Média")] + themes, 12):
        for c, value in enumerate(row, 6):
            ws.cell(r, c, value)
    style_header(ws[12][5:8])
    style_table(ws, 12, 16, 6, 8)

    ws["A19"] = "Fluxo SharePoint/OneDrive"
    style_section(ws["A19"])
    ws.merge_cells("A19:H19")
    flow = [
        "1. Subir este arquivo no SharePoint/OneDrive da empresa.",
        "2. Manter a aba Base automática como destino da integração.",
        "3. A plataforma NPS grava respostas no Supabase e atualiza a planilha via Microsoft Graph API.",
        "4. O time pode criar análises manuais em novas abas sem alterar a Base automática.",
        "5. Heatmaps e replays continuam no Microsoft Clarity; a planilha recebe apenas link opcional quando houver análise manual.",
    ]
    for idx, text in enumerate(flow, 20):
        ws.cell(idx, 1, text)
        ws.merge_cells(start_row=idx, start_column=1, end_row=idx, end_column=8)
        ws.cell(idx, 1).font = Font(name="Calibri", size=10, color=TEXT)
        ws.cell(idx, 1).alignment = Alignment(wrap_text=True)

    chart = BarChart()
    chart.title = "Distribuição NPS"
    chart.y_axis.title = "Quantidade"
    chart.x_axis.title = "Grupo"
    chart.add_data(Reference(ws, min_col=3, min_row=12, max_row=15), titles_from_data=True)
    chart.set_categories(Reference(ws, min_col=1, min_row=13, max_row=15))
    chart.height = 7
    chart.width = 12
    ws.add_chart(chart, "A27")

    set_widths(ws, {"A": 22, "B": 20, "C": 16, "D": 14, "E": 4, "F": 32, "G": 14, "H": 14})
    for row in range(1, 35):
        ws.row_dimensions[row].height = 24


def build_config(ws):
    ws.sheet_view.showGridLines = False
    ws["A1"] = "Configurações da integração"
    style_title(ws["A1"], 20)
    rows = [
        ("Item", "Definição recomendada"),
        ("Arquivo", "Respostas NPS 2026"),
        ("Local", "SharePoint/OneDrive corporativo"),
        ("Aba de destino", "Base automática"),
        ("Sistema de origem", "Plataforma NPS Prime Control"),
        ("Banco de origem", "Supabase"),
        ("Integração", "Microsoft Graph API"),
        ("Atualização", "Automática a cada resposta concluída e sob demanda no dashboard"),
        ("Regra de segurança", "Não enviar senha ou dado sensível; usar app autorizado no Microsoft Entra ID"),
        ("Clarity", "Usado para heatmaps e replay; link pode ser registrado manualmente quando necessário"),
    ]
    for r, row in enumerate(rows, 3):
        for c, value in enumerate(row, 1):
            ws.cell(r, c, value)
    style_header(ws[3])
    style_table(ws, 3, 12, 1, 2)
    set_widths(ws, {"A": 28, "B": 92})
    ws.freeze_panes = "A4"


def build_dictionary(ws):
    ws.sheet_view.showGridLines = False
    ws["A1"] = "Dicionário de dados"
    style_title(ws["A1"], 20)
    headers = ("Campo", "Descrição", "Origem", "Editável?")
    for c, value in enumerate(headers, 1):
        ws.cell(3, c, value)
    for r, row in enumerate(DICTIONARY_ROWS, 4):
        for c, value in enumerate(row, 1):
            ws.cell(r, c, value)
    style_header(ws[3])
    style_table(ws, 3, 3 + len(DICTIONARY_ROWS), 1, 4)
    set_widths(ws, {"A": 34, "B": 72, "C": 18, "D": 14})
    ws.freeze_panes = "A4"


if __name__ == "__main__":
    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    create_workbook()

    wb = load_workbook(OUTPUT, data_only=False)
    required_sheets = {"Resumo executivo", "Base automática", "Configurações", "Dicionário"}
    missing = required_sheets - set(wb.sheetnames)
    if missing:
        raise SystemExit(f"Abas ausentes: {sorted(missing)}")
    if wb["Base automática"].max_column != len(BASE_HEADERS):
        raise SystemExit("Quantidade de colunas da Base automática não confere.")
    print(OUTPUT.resolve())
